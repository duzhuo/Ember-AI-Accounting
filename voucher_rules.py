"""Config-driven voucher rule engine."""

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from string import Formatter
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from voucher_models import SalesTransaction, Voucher, VoucherLine


DEFAULT_RULE_CONFIG_PATH = Path("data/config/voucher_rules.xlsx")


@dataclass(frozen=True)
class VoucherRuleLine:
    """One configurable accounting entry template."""

    rule_code: str
    business_type: str
    product_type: str
    tax_rate: str
    document_type: str
    line_no: int
    debit_credit: str
    account_code: str
    account_name: str
    amount_field: str
    customer_source: str
    tax_code_rule: str
    profit_center_source: str
    cost_center_source: str
    assignment_source: str
    text_template: str


RULE_HEADERS = [
    "rule_code",
    "business_type",
    "product_type",
    "tax_rate",
    "document_type",
    "line_no",
    "debit_credit",
    "account_code",
    "account_name",
    "amount_field",
    "customer_source",
    "tax_code_rule",
    "profit_center_source",
    "cost_center_source",
    "assignment_source",
    "text_template",
]


DEFAULT_SALES_REVENUE_RULES = [
    [
        "SALES_REVENUE_STANDARD",
        "sales_revenue",
        "*",
        "*",
        "DR",
        1,
        "S",
        "112200",
        "应收账款",
        "total_amount",
        "customer",
        "",
        "profit_center",
        "",
        "contract_no",
        "确认客户{customer_name}销售收入，发票{invoice_no}",
    ],
    [
        "SALES_REVENUE_STANDARD",
        "sales_revenue",
        "software|service|saas",
        "*",
        "DR",
        2,
        "H",
        "600101",
        "主营业务收入-软件服务",
        "tax_excluded_amount",
        "",
        "by_tax_rate",
        "profit_center",
        "cost_center",
        "contract_no",
        "确认客户{customer_name}销售收入，发票{invoice_no}",
    ],
    [
        "SALES_REVENUE_STANDARD",
        "sales_revenue",
        "goods",
        "*",
        "DR",
        2,
        "H",
        "600102",
        "主营业务收入-商品销售",
        "tax_excluded_amount",
        "",
        "by_tax_rate",
        "profit_center",
        "cost_center",
        "contract_no",
        "确认客户{customer_name}销售收入，发票{invoice_no}",
    ],
    [
        "SALES_REVENUE_STANDARD",
        "sales_revenue",
        "*",
        "*",
        "DR",
        3,
        "H",
        "22210105",
        "应交税费-应交增值税-销项税额",
        "tax_amount",
        "",
        "by_tax_rate",
        "profit_center",
        "",
        "contract_no",
        "确认客户{customer_name}销售收入，发票{invoice_no}",
    ],
]


def money(value: Decimal) -> Decimal:
    """Round money to two decimal places."""

    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def ensure_default_rule_config(path: str | Path = DEFAULT_RULE_CONFIG_PATH) -> Path:
    """Create a default Excel rule table if it does not exist."""

    rule_path = Path(path)
    if rule_path.exists():
        return rule_path

    rule_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "voucher_rules"
    sheet.append(RULE_HEADERS)
    for row in DEFAULT_SALES_REVENUE_RULES:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            38,
        )

    table_ref = f"A1:P{len(DEFAULT_SALES_REVENUE_RULES) + 1}"
    table = Table(displayName="VoucherRules", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    workbook.save(rule_path)
    return rule_path


def load_voucher_rule_lines(path: str | Path = DEFAULT_RULE_CONFIG_PATH) -> list[VoucherRuleLine]:
    """Load voucher rule lines from an Excel configuration table."""

    rule_path = ensure_default_rule_config(path)
    workbook = load_workbook(rule_path, data_only=True)
    worksheet = workbook["voucher_rules"] if "voucher_rules" in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Rule config is empty: {rule_path}")

    headers = [_normalize_header(header) for header in rows[0]]
    missing = sorted(set(RULE_HEADERS) - set(headers))
    if missing:
        raise ValueError(f"Rule config missing columns: {', '.join(missing)}")

    rule_lines: list[VoucherRuleLine] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        values = dict(zip(headers, row))
        try:
            rule_lines.append(
                VoucherRuleLine(
                    rule_code=_string(values["rule_code"]),
                    business_type=_string(values["business_type"]),
                    product_type=_string(values["product_type"], "*"),
                    tax_rate=_string(values["tax_rate"], "*"),
                    document_type=_string(values["document_type"], "DR"),
                    line_no=int(values["line_no"]),
                    debit_credit=_string(values["debit_credit"]),
                    account_code=_string(values["account_code"]),
                    account_name=_string(values["account_name"]),
                    amount_field=_string(values["amount_field"]),
                    customer_source=_string(values["customer_source"]),
                    tax_code_rule=_string(values["tax_code_rule"]),
                    profit_center_source=_string(values["profit_center_source"]),
                    cost_center_source=_string(values["cost_center_source"]),
                    assignment_source=_string(values["assignment_source"]),
                    text_template=_string(values["text_template"]),
                ),
            )
        except Exception as exc:
            raise ValueError(f"Invalid rule config at row {row_no}: {exc}") from exc

    return rule_lines


def build_sales_revenue_voucher(
    txn: SalesTransaction,
    rule_config_path: str | Path = DEFAULT_RULE_CONFIG_PATH,
    rule_lines: list[VoucherRuleLine] | None = None,
) -> Voucher:
    """Generate a voucher draft from configurable sales revenue rules."""

    warnings: list[str] = []
    calculated_total = money(txn.tax_excluded_amount + txn.tax_amount)
    if calculated_total != money(txn.total_amount):
        warnings.append(
            "Total amount does not equal tax excluded amount plus tax amount.",
        )

    if rule_lines is None:
        rule_lines = load_voucher_rule_lines(rule_config_path)

    matching_rules = _match_rule_lines(
        rule_lines,
        business_type="sales_revenue",
        product_type=txn.product_type,
        tax_rate=txn.tax_rate,
    )
    if not matching_rules:
        raise ValueError("No voucher rules matched sales revenue transaction.")

    summary = _render_template(matching_rules[0].text_template, txn)
    lines = [
        _build_voucher_line(rule_line, txn, summary)
        for rule_line in sorted(matching_rules, key=lambda item: item.line_no)
    ]

    debit_total = sum(line.amount for line in lines if line.debit_credit == "S")
    credit_total = sum(line.amount for line in lines if line.debit_credit == "H")
    if money(debit_total) != money(credit_total):
        warnings.append("Voucher is not balanced.")

    return Voucher(
        voucher_id=f"VR-{txn.transaction_id}",
        company_code=txn.company_code,
        document_type=matching_rules[0].document_type,
        document_date=txn.document_date,
        posting_date=txn.posting_date,
        reference=txn.invoice_no,
        header_text=summary,
        source_transaction_id=txn.transaction_id,
        confidence=Decimal("0.95") if not warnings else Decimal("0.70"),
        warnings=warnings,
        lines=lines,
    )


def _build_voucher_line(
    rule_line: VoucherRuleLine,
    txn: SalesTransaction,
    fallback_text: str,
) -> VoucherLine:
    customer_code = txn.customer_code if rule_line.customer_source == "customer" else ""
    customer_name = txn.customer_name if rule_line.customer_source == "customer" else ""

    return VoucherLine(
        line_no=rule_line.line_no,
        debit_credit=rule_line.debit_credit,
        account_code=rule_line.account_code,
        account_name=rule_line.account_name,
        amount=money(_get_decimal_field(txn, rule_line.amount_field)),
        currency=txn.currency,
        customer_code=customer_code,
        customer_name=customer_name,
        tax_code=_resolve_tax_code(rule_line.tax_code_rule, txn),
        profit_center=_get_string_field(txn, rule_line.profit_center_source),
        cost_center=_get_string_field(txn, rule_line.cost_center_source),
        assignment=_get_string_field(txn, rule_line.assignment_source),
        text=_render_template(rule_line.text_template, txn) or fallback_text,
    )


def _match_rule_lines(
    rule_lines: list[VoucherRuleLine],
    business_type: str,
    product_type: str,
    tax_rate: Decimal,
) -> list[VoucherRuleLine]:
    matched = []
    for rule_line in rule_lines:
        if rule_line.business_type != business_type:
            continue
        if not _matches_token(rule_line.product_type, product_type):
            continue
        if not _matches_tax_rate(rule_line.tax_rate, tax_rate):
            continue
        matched.append(rule_line)
    return matched


def _matches_token(pattern: str, value: str) -> bool:
    if pattern in {"", "*"}:
        return True
    choices = {item.strip().lower() for item in pattern.split("|")}
    return value.strip().lower() in choices


def _matches_tax_rate(pattern: str, value: Decimal) -> bool:
    if pattern in {"", "*"}:
        return True
    return money(Decimal(pattern) * Decimal("100")) == money(value * Decimal("100"))


def _resolve_tax_code(rule: str, txn: SalesTransaction) -> str:
    if rule == "by_tax_rate":
        rate = money(txn.tax_rate * Decimal("100"))
        if rate == Decimal("13.00"):
            return "X1"
        if rate == Decimal("6.00"):
            return "X6"
        if rate == Decimal("0.00"):
            return "X0"
        return "X?"
    return rule


def _get_decimal_field(txn: SalesTransaction, field_name: str) -> Decimal:
    value = getattr(txn, field_name, None)
    if not isinstance(value, Decimal):
        raise ValueError(f"Amount field is not a Decimal field: {field_name}")
    return value


def _get_string_field(txn: SalesTransaction, field_name: str) -> str:
    if not field_name:
        return ""
    value = getattr(txn, field_name, "")
    return str(value or "")


def _render_template(template: str, txn: SalesTransaction) -> str:
    if not template:
        return ""
    allowed_fields = {field for _, field, _, _ in Formatter().parse(template) if field}
    values: dict[str, Any] = {field: getattr(txn, field, "") for field in allowed_fields}
    return template.format(**values)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _string(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()
